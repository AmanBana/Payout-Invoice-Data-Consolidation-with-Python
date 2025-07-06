#!/usr/bin/env python
# coding: utf-8

# ## 1st Task

# In[6]:


# Task 1: Extracting specific cell data from 'Summary' sheet across multiple Excel files
import os
from openpyxl import load_workbook
import pandas as pd

# PATHS
input_folder = r"C:\Users\barbi\Desktop\Financial Data Consolidation Hub\Payout Summary & Order Level Sales"
target_file = r"C:\Users\barbi\Desktop\Financial Data Consolidation Hub\consolidated_file.xlsx"


records = []

for file in os.listdir(input_folder):
    if file.endswith('.xlsx') or file.endswith('.xls'):
        file_path = os.path.join(input_folder, file)

        try:
            wb = load_workbook(filename=file_path, data_only=True)
            if 'Summary' in wb.sheetnames:
                ws = wb['Summary']

                row = {
                    "Brand": ws['B5'].value,
                    "Location": ws['B6'].value,
                    "City": ws['B7'].value,
                    "Res-Id": ws['B8'].value,
                    "Payout Period": ws['C12'].value,
                    "Payout Settlement Date": ws['C13'].value,
                    "Total Payout": ws['C14'].value,
                    "Total Orders (Delivered + Cancelled)": ws['C15'].value,
                    "Bank UTR": ws['C16'].value,
                    "File Name": file
                }
                records.append(row)
            else:
                print(f"'Summary' sheet not found in {file}")
        except Exception as e:
            print(f"Could not process {file} due to: {e}")
    else:
        print(f"Skipping unsupported file format: {file}")

# Write to Excel
if records:
    df = pd.DataFrame(records)
    # This part inherently creates the file and the "Summary" sheet with headers if they don't exist or overwrites them if they do
    with pd.ExcelWriter(target_file, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name='Summary', index=False)
    print(f"Task 1: Summary data successfully saved to {target_file}")
else:
    print("Task 1: No summary data extracted.")


# ## 2nd Task

# In[9]:


# Task 2: Extract Payout Break Up for all files and add brand summary

# Load the target workbook in read/write mode
target_wb = load_workbook(target_file)

# Sheet name for Task 2's output
sheet_name_task2 = "Payout Breakup Tab"

# Create sheet and add headers if it does not exist
if sheet_name_task2 not in target_wb.sheetnames:
    target_ws = target_wb.create_sheet(sheet_name_task2)
    headers = ["SR.No", "Sub-Category", "Particulars", "Delivered Orders", "Cancelled Orders", "Total", "Brand", "Res-Id", "Payout Period", "File Name"]
    target_ws.append(headers)
else:
    # Select existing sheet
    target_ws = target_wb[sheet_name_task2]

# Load Summary Sheet for file-to-brand mapping
summary_df = pd.read_excel(target_file, sheet_name="Summary")

# Get all relevant Excel files in the input folder
file_list = [f for f in os.listdir(input_folder) if f.lower().endswith((".xlsx", ".xls"))]

# Determine the starting SR.No for continuous numbering
if target_ws.max_row == 1 and target_ws.cell(row=1, column=1).value == "SR.No":
    sr_no = 1
elif target_ws.max_row > 1:
    sr_no = target_ws.max_row + 1
else:
    sr_no = 1

# Process each restaurant file
for file in file_list:
    file_path = os.path.join(input_folder, file)

    try:
        # Load the "Payout Breakup" sheet from the source file
        df = pd.read_excel(file_path, sheet_name="Payout Breakup")

        # Extract specific data range
        extracted_data = df.iloc[2:33, 1:6]

        # Get Brand, Res-Id, and Payout Period from Summary sheet
        summary_data = summary_df.loc[summary_df["File Name"] == file, ["Brand", "Res-Id", "Payout Period"]].values[0]
        brand, res_id, payout_period = summary_data

        # Append data to the target sheet
        for _, row in extracted_data.iterrows():
            formatted_row = [sr_no] + row.tolist() + [brand, res_id, payout_period, file]
            target_ws.append(formatted_row)
            sr_no += 1

    except Exception as e:
        print(f"Task 2: Error processing file {file} due to: {e}")

# Save the updated workbook
target_wb.save(target_file)
print(f"Task 2: Payout Breakup data successfully appended to '{sheet_name_task2}' in {target_file}")


# ## 3rd Task

# In[14]:


# Task 3: Extract Order Level Details for all files and add brand summary

# Load the target workbook in read/write mode
target_wb = load_workbook(target_file)

sheet_name_task3 = "Order Level"

if sheet_name_task3 not in target_wb.sheetnames:
    target_ws = target_wb.create_sheet(sheet_name_task3)

    source_order_level_headers = []
    # Find the first Excel file in the input folder
    sample_file_path = None
    for f in os.listdir(input_folder):
        if f.lower().endswith((".xlsx", ".xls")):
            sample_file_path = os.path.join(input_folder, f)
            break

    if sample_file_path:
        try:
            # Read the "Order Level" sheet from the sample file, explicitly setting header to row 3 (index 2).
            temp_df = pd.read_excel(sample_file_path, sheet_name="Order Level", header=2)
            source_order_level_headers = list(temp_df.columns)
        except Exception as e:
            # Log error if unable to read sample file for header determination.
            print(f"Task 3: Could not read 'Order Level' sheet from {sample_file_path} to determine headers. Error: {e}")
            pass # Continue execution as per project guidelines.
    
   
    all_target_headers_task3 = ["Brand", "Res-Id", "Payout Period", "File Name"] + source_order_level_headers
    target_ws.append(all_target_headers_task3) # Add headers to the newly created sheet
else:
    
    target_ws = target_wb[sheet_name_task3]

# Load the Summary Sheet from the target file for file-to-brand mapping
summary_df = pd.read_excel(target_file, sheet_name="Summary")

# Get all relevant Excel files in the input folder
file_list = [f for f in os.listdir(input_folder) if f.lower().endswith((".xlsx", ".xls"))]

# Process each restaurant file
for file in file_list:
    file_path = os.path.join(input_folder, file)

    try:
        df = pd.read_excel(file_path, sheet_name="Order Level")

        # Extract rows from row 4 onwards (index 3 onwards), taking all columns.
        extracted_data = df.iloc[3:, :]

        # Get Brand, Res-Id, and Payout Period from Summary sheet based on File Name
        summary_data = summary_df.loc[summary_df["File Name"] == file, ["Brand", "Res-Id", "Payout Period"]].values[0]
        brand, res_id, payout_period = summary_data

        # Append data with metadata to the target sheet
        for _, row in extracted_data.iterrows():
            formatted_row = [brand, res_id, payout_period, file] + row.tolist()
            target_ws.append(formatted_row)

    except Exception as e:
        print(f"Task 3: Error processing file {file} due to: {e}")

# Save the updated workbook
target_wb.save(target_file)
print(f"Task 3: Order Level data successfully appended to '{sheet_name_task3}' in {target_file}")


# ## Task 4

# In[33]:


# Task 4: Extract Details from Commission Invoice PDFs and fill "Commission Invoice" tab
import pdfplumber
import pandas as pd
import os
import re
from datetime import datetime

# setting up Path
pdf_folder = r"C:\Users\barbi\Desktop\Financial Data Consolidation Hub\Commission Invoices" # Updated to the correct path you provided
excel_file = r"C:\Users\barbi\Desktop\Financial Data Consolidation Hub\consolidated_file.xlsx" # Our standard target file
sheet_name = "Commission Invoice"

# Loading existing Excel data
df_existing = pd.read_excel(excel_file, sheet_name=sheet_name)

# Helper to extract a value using regex
def extract_value(pattern, text, default=""):
    match = re.search(pattern, text)
    return match.group(1).strip() if match else default

# Parse a single PDF file
def parse_pdf(file_path):
    text = ""
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() + "\n"

    filename = os.path.basename(file_path)

    # Extract values
    invoice_date_raw = extract_value(r"Invoice Date\s*:\s*(.+)", text)
    invoice_date = datetime.strptime(invoice_date_raw.split()[0], "%Y-%m-%d").date()
    year = invoice_date.year
    month = str(invoice_date.month).zfill(2)
    fy_year = f"{year}-{str(year+1)[-2:]}" if month >= "04" else f"{year-1}-{str(year)[-2:]}"
    
    data = {
        "payout_period": extract_value(r"Service Period\s*:\s*(.+)", text),
        "file_name": filename,
        "fy_year": fy_year,
        "year": year,
        "month": month,
        "irn": extract_value(r"IRN\s*:\s*(\w+)", text),
        "mann_gstin": extract_value(r"GSTIN\s*:\s*(29ABNFM9601R1Z9)", text),
        "swiggy_gstin": extract_value(r"GSTIN\s*:\s*(29AAFCB7707D1ZQ)", text),
        "pan": extract_value(r"PAN\s*:\s*([A-Z0-9]+)", text),
        "invoice_date": invoice_date_raw,
        "invoice_number": extract_value(r"Invoice Number\s*:\s*(\w+)", text),
        "original_invoice_number": extract_value(r"Original Invoice\s*No:\s*(.*)", text),
        "invoice_type": extract_value(r"Invoice Type\s*:\s*(\w+)", text),
        "brand_id": extract_value(r"Restaurant / Store ID\s*:\s*(\d+)", text),
        "other_charges_reimbursement_of_discount": extract_value(r"Other Charges - Reimbursement\s*of Discount\s*([\d.,]+)", text),
        "grand_total": extract_value(r"Grand Total\s*([\d.,]+)", text)
    }

    # Extracting line item details (for now assuming 2 items max)
    item_lines = re.findall(r"(\d+)\s+([^\n]+?)\s+996211\s+OTH\s+1\s+([\d.,]+)\s+([\d.,]+)\s+0\s+([\d.,]+)\s+9\s+([\d.,]+)\s+9\s+([\d.,]+).*?([\d.,]+)", text)
    
    rows = []
    for item in item_lines:
        row = data.copy()
        row.update({
            "sr_no": item[0],
            "description": item[1],
            "hsn": "996211",
            "unit_of_measure": "OTH",
            "quantity": "1",
            "unit_price": item[2],
            "base_amount": item[3],
            "discount": "0",
            "assessable_value": item[4],
            "cgst_rate": "9",
            "cgst_amount": item[5],
            "sgst_rate": "9",
            "sgst_amount": item[6],
            "igst_rate": "0",
            "igst_amount": "0",
            "comp_cess_rate": "0",
            "comp_cess_amount": "0",
            "state_cess_rate": "0",
            "state_cess_amount": "0",
            "total_amount": item[7],
        })
        rows.append(row)

    return rows

# Processing all PDFs
all_rows = []
for filename in os.listdir(pdf_folder):
    if filename.lower().endswith(".pdf"):
        file_path = os.path.join(pdf_folder, filename)
        try:
            rows = parse_pdf(file_path)
            all_rows.extend(rows)
            print(f"Parsed: {filename}")
        except Exception as e:
            print(f"Failed to parse {filename}: {e}")

# Combine and append
df_new = pd.DataFrame(all_rows)
df_final = pd.concat([df_existing, df_new], ignore_index=True)

# Save to Excel
with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_final.to_excel(writer, sheet_name=sheet_name, index=False)

print("Data extraction and appending completed")


# In[ ]:




